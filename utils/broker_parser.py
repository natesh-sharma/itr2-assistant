from datetime import datetime
from typing import Any, BinaryIO, Dict, List, Optional, Union

import openpyxl
import pandas as pd


def parse_zerodha_pnl(file: Union[str, BinaryIO]) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(file, data_only=True)

    equity_stcg: List[Dict] = []
    equity_ltcg: List[Dict] = []
    mf_ltcg_equity: List[Dict] = []
    mf_ltcg_debt: List[Dict] = []
    dividends: List[Dict] = []

    tradewise_sheet = _find_sheet(wb, "Tradewise Exits")
    if tradewise_sheet:
        sections = _split_sections(tradewise_sheet)
        for section_name, rows in sections.items():
            trades = _parse_trade_rows(rows)
            name_lower = section_name.lower()
            if "mutual fund" in name_lower:
                for t in trades:
                    scheme = t["symbol"].lower()
                    if any(kw in scheme for kw in ["liquid", "debt", "money market", "gilt", "bond", "overnight", "ultra short"]):
                        mf_ltcg_debt.append(t)
                    else:
                        mf_ltcg_equity.append(t)
            elif "short term" in name_lower:
                equity_stcg.extend(trades)
            elif "long term" in name_lower:
                equity_ltcg.extend(trades)
            elif "intraday" in name_lower or "speculative" in name_lower:
                equity_stcg.extend(trades)

    div_sheet = _find_sheet(wb, "Dividends")
    if div_sheet:
        dividends = _parse_dividend_sheet(div_sheet)

    stcg_total = sum(t.get("profit", 0) for t in equity_stcg)
    ltcg_equity_total = sum(t.get("profit", 0) for t in equity_ltcg) + sum(t.get("profit", 0) for t in mf_ltcg_equity)
    ltcg_debt_total = sum(t.get("profit", 0) for t in mf_ltcg_debt)
    dividend_total = sum(d.get("amount", 0) for d in dividends)

    return {
        "equity_stcg": equity_stcg,
        "equity_ltcg": equity_ltcg,
        "mf_ltcg_equity": mf_ltcg_equity,
        "mf_ltcg_debt": mf_ltcg_debt,
        "dividends": dividends,
        "summary": {
            "stcg_total": round(stcg_total, 2),
            "ltcg_equity_total": round(ltcg_equity_total, 2),
            "ltcg_debt_total": round(ltcg_debt_total, 2),
            "dividend_total": round(dividend_total, 2),
        },
    }


def _find_sheet(wb, keyword: str):
    for name in wb.sheetnames:
        if keyword.lower() in name.lower():
            return wb[name]
    return None


def _split_sections(ws) -> Dict[str, List[List]]:
    sections: Dict[str, List[List]] = {}
    current_section = None
    header_row = None

    for row in ws.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        non_none = [v for v in row if v is not None and str(v).strip()]
        if len(non_none) == 1:
            label = str(non_none[0]).strip()
            section_markers = ["equity - intraday", "equity - short term", "equity - long term",
                               "equity - buyback", "mutual funds", "f&o", "currency", "commodity"]
            if any(m in label.lower() for m in section_markers):
                current_section = label
                header_row = None
                sections[current_section] = []
                continue

        if current_section is not None:
            if header_row is None:
                str_vals = [str(v).lower().strip() for v in row if v is not None]
                if "symbol" in str_vals:
                    header_row = [str(v).strip() if v is not None else "" for v in row]
                    continue
            else:
                cell1 = row[1] if len(row) > 1 else None
                if cell1 is not None and str(cell1).strip() and str(cell1).strip() not in ("Symbol", ""):
                    sections[current_section].append((header_row, list(row)))

    return sections


def _parse_trade_rows(rows: List) -> List[Dict]:
    trades = []
    if not rows:
        return trades

    for header_row, data_row in rows:
        try:
            col_idx = {h.lower(): i for i, h in enumerate(header_row) if h}

            symbol_i = _find_idx(col_idx, ["symbol", "scrip"])
            isin_i = _find_idx(col_idx, ["isin"])
            buy_date_i = _find_idx(col_idx, ["entry date", "buy date"])
            sell_date_i = _find_idx(col_idx, ["exit date", "sell date"])
            qty_i = _find_idx(col_idx, ["quantity", "qty"])
            buy_val_i = _find_idx(col_idx, ["buy value", "buy amount"])
            sell_val_i = _find_idx(col_idx, ["sell value", "sale value"])
            profit_i = _find_idx(col_idx, ["profit", "p&l", "realized p&l"])
            fmv_i = _find_idx(col_idx, ["fair market value", "fmv"])
            holding_i = _find_idx(col_idx, ["period of holding"])

            symbol = _safe_str(data_row, symbol_i)
            if not symbol:
                continue

            qty = _safe_float(data_row, qty_i)
            if qty is None or qty == 0:
                continue

            trade: Dict[str, Any] = {
                "symbol": symbol,
                "isin": _safe_str(data_row, isin_i),
                "buy_date": _safe_date(data_row, buy_date_i),
                "sell_date": _safe_date(data_row, sell_date_i),
                "qty": qty,
                "buy_value": _safe_float(data_row, buy_val_i) or 0,
                "sell_value": _safe_float(data_row, sell_val_i) or 0,
                "profit": _safe_float(data_row, profit_i) or 0,
                "holding_days": int(_safe_float(data_row, holding_i) or 0),
            }

            fmv = _safe_float(data_row, fmv_i)
            if fmv is not None:
                trade["fmv"] = fmv

            trades.append(trade)
        except Exception:
            continue

    return trades


def _parse_dividend_sheet(ws) -> List[Dict]:
    divs = []
    in_section = False
    headers = None

    for row in ws.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        str_vals = [str(v).lower().strip() for v in row if v is not None]

        if "symbol" in str_vals and any("dividend" in v or "amount" in v for v in str_vals):
            headers = [str(v).strip().lower() if v is not None else "" for v in row]
            in_section = True
            continue

        if in_section and headers and row[1] is not None:
            try:
                col_idx = {h: i for i, h in enumerate(headers) if h}
                symbol_i = _find_idx(col_idx, ["symbol", "scrip"])
                amount_i = _find_idx(col_idx, ["net dividend amount", "dividend amount", "amount"])
                date_i = _find_idx(col_idx, ["ex-date", "ex date", "date"])

                amount = _safe_float(row, amount_i)
                if amount is None or amount == 0:
                    continue

                divs.append({
                    "symbol": _safe_str(row, symbol_i),
                    "amount": round(amount, 2),
                    "ex_date": _safe_date(row, date_i),
                })
            except Exception:
                continue

    return divs


def _find_idx(col_idx: Dict[str, int], candidates: List[str]) -> Optional[int]:
    for c in candidates:
        for key, idx in col_idx.items():
            if c in key:
                return idx
    return None


def _safe_str(row: list, idx: Optional[int]) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _safe_float(row: list, idx: Optional[int]) -> Optional[float]:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        try:
            return float(str(val).replace(",", "").replace("₹", "").strip())
        except (ValueError, TypeError):
            return None


def _safe_date(row: list, idx: Optional[int]) -> Optional[str]:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s
